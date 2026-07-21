"""Import the client's "Drawings Number Book.xlsx" as the seed registry.

Usage: python scripts/import_book.py "path/to/Drawings Number Book.xlsx"

- Every category tab (all except 'Main Book') becomes a project (source
  'book_import'); its rows become drawings under that project.
- 'Main Book' rows are imported afterwards WITHOUT a project, skipping DWG
  numbers already imported from a category tab (the Main Book is the master
  list; tabs are curated subsets).
- Set # values become drawing_sets rows (per project + set number) and
  drawings are linked to them.
- Rows sharing the same normalized DWG number are linked into one version
  group (versions identified by DWG number + date, per the requirements).
- Messy source data is preserved raw (contract notes, date ranges); a
  best-effort year is parsed for version ordering.

Idempotent-ish: running twice would duplicate rows - wipe book_import rows
first with --replace.
"""
import sys

import openpyxl

sys.path.insert(0, ".")

from app.db import pool  # noqa: E402
from app.repositories import DrawingRepository, ProjectRepository  # noqa: E402
from app.services import matching  # noqa: E402

MAIN = "Main Book"


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def import_book(path: str, replace: bool) -> None:
    pool.open()
    projects = ProjectRepository(pool)
    drawings = DrawingRepository(pool)

    if replace:
        with pool.connection() as conn:
            conn.execute("DELETE FROM drawings WHERE source = 'book_import'")
            conn.execute(
                "DELETE FROM drawing_sets WHERE id NOT IN "
                "(SELECT DISTINCT set_id FROM drawings WHERE set_id IS NOT NULL)"
            )
            conn.execute("DELETE FROM projects WHERE source = 'book_import'")
        print("cleared previous book_import rows")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    seen_norms: dict[str, str] = {}  # norm -> first drawing_id (for version linking)
    set_cache: dict[tuple, str] = {}  # (project_id, set_number) -> set_id
    stats = {"projects": 0, "drawings": 0, "sets": 0, "version_links": 0, "skipped": 0}

    def import_row(row: tuple, project_id: str | None) -> None:
        dwg_raw, sheets, description, contract, date, set_no = (
            clean(row[0]), row[1], clean(row[2]), clean(row[3]), row[4],
            clean(row[5]) if len(row) > 5 else None,
        )
        if dwg_raw is None and description is None:
            return
        if dwg_raw in (None, "DWG #"):  # header/continuation rows
            stats["skipped"] += 1
            return

        parsed = matching.parse_filename(dwg_raw)
        norm = parsed["dwg_candidates"][0]["norm"] if parsed["dwg_candidates"] else None

        # Main Book pass: skip rows already imported from a category tab
        if project_id is None and norm and norm in seen_norms:
            stats["skipped"] += 1
            return

        set_id = None
        if set_no:
            key = (project_id, set_no)
            if key not in set_cache:
                created = drawings.create_set(project_id, set_no, None)
                set_cache[key] = created["set_id"]
                stats["sets"] += 1
            set_id = set_cache[key]

        try:
            sheet_count = int(sheets) if sheets is not None else None
        except (TypeError, ValueError):
            sheet_count = None

        created = drawings.create(
            {
                "project_id": project_id,
                "set_id": set_id,
                "dwg_number": dwg_raw,
                "dwg_number_norm": norm,
                "description": description,
                "contract_number": contract,
                "drawing_date": clean(date),
                "year": matching.parse_year(date),
                "sheet_count": sheet_count,
                "source": "book_import",
            }
        )
        stats["drawings"] += 1

        if norm:
            if norm in seen_norms:
                # same DWG number recorded more than once -> versions of one drawing
                drawings.link_versions(seen_norms[norm], created["drawing_id"])
                stats["version_links"] += 1
            else:
                seen_norms[norm] = created["drawing_id"]

    # 1) category tabs -> projects with drawings
    for sheet_name in wb.sheetnames:
        if sheet_name == MAIN:
            continue
        project = projects.create(sheet_name, None, None, source="book_import")
        stats["projects"] += 1
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3, values_only=True):
            import_row(row, project["project_id"])

    # 2) Main Book -> remaining drawings, no project
    ws = wb[MAIN]
    for row in ws.iter_rows(min_row=3, values_only=True):
        import_row(row, None)

    pool.close()
    print(stats)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("usage: python scripts/import_book.py <xlsx path> [--replace]")
    import_book(sys.argv[1], "--replace" in sys.argv)
